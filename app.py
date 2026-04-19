import os
import re
import io
import pandas as pd
from datetime import datetime, date
from flask import Flask, request, render_template, send_file, jsonify
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB limit

ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# ── Mobile Cleaning ──────────────────────────────────────────────────────────

def clean_mobile(val):
    if pd.isna(val) or str(val).strip() == '':
        return '9999999990'
    digits = re.sub(r'\D', '', str(val))
    if len(digits) == 14 and digits.startswith('0091'):
        digits = digits[4:]
    elif len(digits) == 13 and digits.startswith('091'):
        digits = digits[3:]
    elif len(digits) == 12 and digits.startswith('91'):
        digits = digits[2:]
    elif len(digits) == 11 and digits.startswith('0'):
        digits = digits[1:]
    if len(digits) == 10:
        return digits
    return '9999999990'


# ── DOB Cleaning ─────────────────────────────────────────────────────────────

def clean_dob(val):
    if pd.isna(val) or str(val).strip() == '':
        return ''
    s = str(val).strip()
    # Excel serial date (numeric string)
    try:
        serial = float(s)
        if 1000 < serial < 100000:
            base = date(1899, 12, 30)
            from datetime import timedelta
            d = base + timedelta(days=int(serial))
            return d.strftime('%Y-%m-%d')
    except ValueError:
        pass
    formats = [
        '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y', '%d.%m.%Y',
        '%Y/%m/%d', '%Y-%m-%d', '%d-%b-%Y', '%d %b %Y',
        '%B %d, %Y', '%d/%m/%y', '%d-%m-%y'
    ]
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return ''


# ── Age Cleaning ─────────────────────────────────────────────────────────────

def clean_age(val):
    if pd.isna(val) or str(val).strip() == '':
        return val
    s = str(val).strip()
    match = re.search(r'(\d+)\s*[yY]', s)
    if match:
        return match.group(1) + 'y'
    return s.strip()


# ── Name Cleaning ─────────────────────────────────────────────────────────────

def clean_name(first, last):
    parts = []
    for p in [first, last]:
        if not pd.isna(p) and str(p).strip():
            parts.append(str(p).strip())
    name = ' '.join(parts)
    name = re.sub(r'[^\w\s\-\']', '', name)
    name = re.sub(r'\s+', ' ', name).strip()
    return name.title() if name else ''


def clean_single_name(val):
    if pd.isna(val) or str(val).strip() == '':
        return ''
    name = re.sub(r'[^\w\s\-\']', '', str(val).strip())
    name = re.sub(r'\s+', ' ', name).strip()
    return name.title()


# ── Column Detection ──────────────────────────────────────────────────────────

def detect_column(columns, keywords):
    """Find the first column name that contains any of the given keywords."""
    cols_lower = {c: re.sub(r'[\s_\n\r\*\.\,\(\)]', '', c.lower()) for c in columns}
    for col, col_norm in cols_lower.items():
        for kw in keywords:
            if kw in col_norm:
                return col
    return None


# ── Core Cleaning Logic ───────────────────────────────────────────────────────

def process_dataframe(df, sheet_name='Sheet1', patient_id_prefix='P'):
    stats = {
        'sheet': sheet_name,
        'input_rows': len(df),
        'empty_rows_removed': 0,
        'junk_removed': 0,
        'duplicates_removed': 0,
        'mobile_defaulted': 0,
        'name_defaulted': 0,
        'dob_blank': 0,
        'malformed_ids': 0,
        'output_rows': 0,
    }

    # Clean column names — drop NaN/float headers, strip whitespace
    df.columns = [str(c).strip() if not pd.isna(c) else f'_col_{i}' for i, c in enumerate(df.columns)]
    # Drop unnamed filler columns (e.g. 'Unnamed: 5', '_col_3')
    df = df.loc[:, ~df.columns.str.match(r'^(Unnamed:|_col_\d+)')]

    # Drop fully blank rows
    before = len(df)
    df = df.dropna(how='all')
    df = df[df.apply(lambda r: r.astype(str).str.strip().ne('').any(), axis=1)]
    stats['empty_rows_removed'] = before - len(df)

    if df.empty:
        stats['output_rows'] = 0
        return pd.DataFrame(), stats

    cols = list(df.columns)

    # ── Map columns ──
    pid_col   = detect_column(cols, ['patientid', 'pid', 'id', 'uhid', 'mrno', 'mrid'])
    fname_col = detect_column(cols, ['firstname', 'fname', 'first'])
    lname_col = detect_column(cols, ['lastname', 'lname', 'last', 'surname'])
    name_col  = detect_column(cols, ['patientname', 'name', 'fullname', 'patient'])
    mobile_col = detect_column(cols, ['mobile', 'phone', 'contact', 'cell', 'number', 'mob'])
    dob_col   = detect_column(cols, ['dob', 'dateofbirth', 'birthdate', 'birth'])
    age_col   = detect_column(cols, ['age'])
    gender_col = detect_column(cols, ['gender', 'sex'])
    address_col = detect_column(cols, ['address', 'addr', 'location', 'city'])

    # ── Build Patient Name ──
    if fname_col and lname_col:
        df['Patient Name'] = df.apply(
            lambda r: clean_name(r.get(fname_col, ''), r.get(lname_col, '')), axis=1
        )
    elif fname_col:
        df['Patient Name'] = df[fname_col].apply(clean_single_name)
    elif name_col:
        df['Patient Name'] = df[name_col].apply(clean_single_name)
    else:
        df['Patient Name'] = ''

    # ── Junk / Test record removal ──
    junk_names = {'test', 'demo', 'dummy', 'sample'}
    def is_junk(name):
        if not name:
            return False
        first_word = name.strip().split()[0].lower() if name.strip() else ''
        return first_word in junk_names
    before = len(df)
    df = df[~df['Patient Name'].apply(is_junk)]
    stats['junk_removed'] = before - len(df)

    # ── Default blank names to ABC ──
    blank_name_mask = df['Patient Name'].str.strip() == ''
    stats['name_defaulted'] = int(blank_name_mask.sum())
    df.loc[blank_name_mask, 'Patient Name'] = 'ABC'

    # ── Mobile ──
    if mobile_col:
        df['Mobile Number'] = df[mobile_col].apply(clean_mobile)
    else:
        df['Mobile Number'] = '9999999990'
    stats['mobile_defaulted'] = int((df['Mobile Number'] == '9999999990').sum())

    # ── DOB ──
    if dob_col:
        df['DOB'] = df[dob_col].apply(clean_dob)
        stats['dob_blank'] = int((df['DOB'] == '').sum())
    else:
        df['DOB'] = ''

    # ── Age ──
    if age_col:
        df['Age'] = df[age_col].apply(clean_age)

    # ── Patient ID ──
    if pid_col:
        df['Patient ID'] = df[pid_col].astype(str).str.strip()
        malformed = df['Patient ID'].apply(
            lambda x: pd.isna(x) or bool(re.search(r'[^a-zA-Z0-9\-_]', str(x))) or str(x) in ('', 'nan', 'NaN')
        )
        stats['malformed_ids'] = int(malformed.sum())
    else:
        prefix = patient_id_prefix if patient_id_prefix else 'P'
        df['Patient ID'] = [f'{prefix}{str(i+1).zfill(4)}' for i in range(len(df))]

    # ── Duplicate removal (name + mobile both match) ──
    df['_name_key'] = df['Patient Name'].str.lower().str.strip()
    df['_mob_key'] = df['Mobile Number']
    before = len(df)
    df = df.drop_duplicates(subset=['_name_key', '_mob_key'], keep='first')
    stats['duplicates_removed'] = before - len(df)
    df = df.drop(columns=['_name_key', '_mob_key'])

    # ── Assemble output columns ──
    out_cols = ['Patient ID', 'Patient Name', 'Mobile Number', 'DOB']
    if age_col:
        out_cols.append('Age')
    if gender_col:
        df['Gender'] = df[gender_col]
        out_cols.append('Gender')
    if address_col:
        df['Address'] = df[address_col]
        out_cols.append('Address')

    df = df[out_cols].reset_index(drop=True)
    stats['output_rows'] = len(df)
    return df, stats


# ── Excel Writer ──────────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill(fill_type='solid', fgColor='1F4E79')
HEADER_FONT   = Font(name='Arial', size=10, bold=True, color='FFFFFF')
ALT_FILL      = PatternFill(fill_type='solid', fgColor='DCE6F1')
NORMAL_FONT   = Font(name='Arial', size=10)
SUMMARY_HEADER_FONT = Font(name='Arial', size=11, bold=True, color='1F4E79')
THIN_BORDER   = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def write_excel(all_dfs, all_stats, original_name, patient_id_prefix=''):
    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Cleaned Patient Data ──
    ws = wb.create_sheet('Cleaned Patient Data')
    if all_dfs:
        combined = pd.concat(all_dfs, ignore_index=True)
    else:
        combined = pd.DataFrame()

    if not combined.empty:
        headers = list(combined.columns)
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for i, row in enumerate(combined.itertuples(index=False), start=2):
            ws.append(list(row))
            fill = ALT_FILL if i % 2 == 0 else PatternFill()
            for cell in ws[i]:
                cell.font = NORMAL_FONT
                cell.fill = fill
                cell.alignment = Alignment(vertical='center')

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = 'A2'

        for col_idx, col_name in enumerate(headers, 1):
            max_len = max(len(str(col_name)), combined[col_name].astype(str).str.len().max())
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    # ── Sheet 2: Summary Report ──
    ws2 = wb.create_sheet('Summary Report')
    ws2.column_dimensions['A'].width = 36
    ws2.column_dimensions['B'].width = 20

    def section(title):
        ws2.append([title])
        ws2[ws2.max_row][0].font = SUMMARY_HEADER_FONT
        ws2.append([])

    def row(label, value):
        ws2.append([label, value])
        for cell in ws2[ws2.max_row]:
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER

    total_input  = sum(s['input_rows']       for s in all_stats)
    total_output = sum(s['output_rows']      for s in all_stats)
    empty_removed = sum(s['empty_rows_removed'] for s in all_stats)
    junk_removed  = sum(s['junk_removed']    for s in all_stats)
    dups_removed  = sum(s['duplicates_removed'] for s in all_stats)
    mob_default   = sum(s['mobile_defaulted'] for s in all_stats)
    name_default  = sum(s['name_defaulted']  for s in all_stats)
    dob_blank     = sum(s['dob_blank']       for s in all_stats)
    bad_ids       = sum(s['malformed_ids']   for s in all_stats)

    ws2.append(['EMR Patient Data Cleaning Report'])
    ws2[ws2.max_row][0].font = Font(name='Arial', size=14, bold=True, color='1F4E79')
    ws2.append([f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'])
    ws2.append([f'Source file: {original_name}'])
    if patient_id_prefix:
        ws2.append([f'Patient ID Prefix used: {patient_id_prefix}'])
    ws2.append([])

    section('Input Data')
    row('Total records received', total_input)
    row('Source sheets processed', len(all_stats))
    ws2.append([])

    section('Cleaning Actions')
    row('Empty rows removed', empty_removed)
    row('Junk/test records removed', junk_removed)
    row('Duplicate records removed', dups_removed)
    row('Names defaulted to ABC', name_default)
    row('Mobiles defaulted to 9999999990', mob_default)
    ws2.append([])

    section('Output Data')
    row('Total cleaned records', total_output)
    row('Records with missing DOB', dob_blank)
    row('Malformed Patient IDs flagged', bad_ids)
    ws2.append([])

    section('Data Quality Applied')
    row('Mobile format', '10-digit, country code stripped')
    row('DOB format', 'YYYY-MM-DD')
    row('Name format', 'Title Case, combined first+last')
    row('Duplicate rule', 'Name + Mobile both match')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Pharmacy: Expiry Date Cleaning ───────────────────────────────────────────

def clean_exp_date(val):
    if pd.isna(val) or str(val).strip() == '':
        return ''
    s = str(val).strip()
    if re.match(r'^[A-Za-z]{3}-\d{4}$', s):
        return s
    formats = [
        '%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y',
        '%m/%d/%Y', '%d.%m.%Y', '%Y/%m/%d', '%d-%b-%Y', '%d %b %Y',
        '%B %d, %Y', '%d/%m/%y', '%d-%m-%y', '%m-%Y', '%m/%Y',
        '%b-%Y', '%b %Y',
    ]
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt).strftime('%b-%Y')
        except ValueError:
            continue
    try:
        return pd.to_datetime(s, dayfirst=True).strftime('%b-%Y')
    except Exception:
        return s


# ── Pharmacy: Pack / Free Pack Parsing ───────────────────────────────────────

def parse_pack(val):
    if pd.isna(val) or str(val).strip() == '':
        return ''
    s = str(val).strip()
    m = re.match(r'^(\d+)\s*\*\s*(\d+)$', s)
    if m:
        return str(int(m.group(1)) + int(m.group(2)))
    m2 = re.match(r'^(\d+)', s)
    if m2:
        return m2.group(1)
    return s


# ── Pharmacy: Item per Pack Parsing ──────────────────────────────────────────

def parse_item_per_pack(val):
    if pd.isna(val) or str(val).strip() == '':
        return ''
    s = str(val).strip()
    m = re.match(r'^(\d+)', s)
    return m.group(1) if m else s


# ── Pharmacy: Core Processing ─────────────────────────────────────────────────

PHARMACY_OUT_COLS = [
    'Supplier Name *', 'Medicine Name*', 'HSN No.', 'Batch No. *',
    'Exp. Date*', 'Pack* & Free Pack', 'Item per Pack',
    'MRP*', 'Pure Cost*', 'Discount *', 'GST No.*',
]

def process_pharmacy_dataframe(df, sheet_name='Sheet1', supplier_name=''):
    stats = {'sheet': sheet_name, 'input_rows': len(df), 'output_rows': 0, 'empty_rows_removed': 0}

    df.columns = [str(c).strip() if not pd.isna(c) else f'_col_{i}' for i, c in enumerate(df.columns)]
    df = df.loc[:, ~df.columns.str.match(r'^(Unnamed:|_col_\d+)')]

    before = len(df)
    df = df.dropna(how='all')
    df = df[df.apply(lambda r: r.astype(str).str.strip().ne('').any(), axis=1)]
    stats['empty_rows_removed'] = before - len(df)

    if df.empty:
        return pd.DataFrame(columns=PHARMACY_OUT_COLS), stats

    cols = list(df.columns)

    med_col      = detect_column(cols, ['medicinename', 'medicine', 'drugname', 'drug', 'itemname', 'productname'])
    hsn_col      = detect_column(cols, ['hsn', 'hsncode', 'hsnno'])
    batch_col    = detect_column(cols, ['batch', 'batchno', 'batchnumber', 'lot', 'lotno'])
    exp_col      = detect_column(cols, ['exp', 'expiry', 'expirydate', 'expdate', 'expdt'])
    pack_col     = detect_column(cols, ['pack', 'qty', 'quantity', 'stock', 'freepack', 'total'])
    ipack_col    = detect_column(cols, ['packing', 'itemperpack', 'perpack', 'packsize', 'unitpack'])
    mrp_col      = detect_column(cols, ['mrp', 'maxretail', 'retailprice', 'listprice'])
    cost_col     = detect_column(cols, ['purecost', 'cost', 'rate', 'purchaserate', 'purchaseprice', 'price'])
    disc_col     = detect_column(cols, ['discount', 'disc', 'dis'])
    gst_col      = detect_column(cols, ['gst', 'gstnumber', 'gstno', 'gstin', 'tax'])

    out = pd.DataFrame(index=range(len(df)))
    out['Supplier Name *']  = supplier_name
    out['Medicine Name*']   = df[med_col].apply(lambda v: str(v).strip().upper() if not pd.isna(v) and str(v).strip() else '') if med_col else ''
    out['HSN No.']          = df[hsn_col].apply(lambda v: str(v).strip() if not pd.isna(v) and str(v).strip() not in ('', 'nan') else '') if hsn_col else ''
    out['Batch No. *']      = df[batch_col].apply(lambda v: str(v).strip() if not pd.isna(v) and str(v).strip() not in ('', 'nan') else '') if batch_col else ''
    out['Exp. Date*']       = df[exp_col].apply(clean_exp_date) if exp_col else ''
    out['Pack* & Free Pack']= df[pack_col].apply(parse_pack) if pack_col else ''
    out['Item per Pack']    = df[ipack_col].apply(parse_item_per_pack) if ipack_col else ''
    out['MRP*']             = df[mrp_col].apply(lambda v: str(v).strip() if not pd.isna(v) and str(v).strip() not in ('', 'nan') else '') if mrp_col else ''
    out['Pure Cost*']       = df[cost_col].apply(lambda v: str(v).strip() if not pd.isna(v) and str(v).strip() not in ('', 'nan') else '') if cost_col else ''
    out['Discount *']       = df[disc_col].apply(lambda v: str(v).strip() if not pd.isna(v) and str(v).strip() not in ('', 'nan') else '') if disc_col else ''
    out['GST No.*']         = df[gst_col].apply(lambda v: str(v).strip() if not pd.isna(v) and str(v).strip() not in ('', 'nan') else '') if gst_col else ''

    out = out[out['Medicine Name*'].str.strip().ne('')]
    out = out.reset_index(drop=True)
    stats['output_rows'] = len(out)
    return out, stats


# ── Pharmacy: Excel Writer ────────────────────────────────────────────────────

def write_pharmacy_excel(all_dfs, all_stats, original_name, supplier_name=''):
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet('Pharmacy Data')
    combined = pd.concat(all_dfs, ignore_index=True) if all_dfs else pd.DataFrame(columns=PHARMACY_OUT_COLS)

    headers = PHARMACY_OUT_COLS
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = PatternFill(fill_type='solid', fgColor='1F4E79')
        cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for i, row_data in enumerate(combined[headers].itertuples(index=False), start=2):
        ws.append(list(row_data))
        fill = PatternFill(fill_type='solid', fgColor='DCE6F1') if i % 2 == 0 else PatternFill()
        for cell in ws[i]:
            cell.font = Font(name='Arial', size=10)
            cell.fill = fill
            cell.alignment = Alignment(vertical='center')

    if not combined.empty:
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = 'A2'
        for col_idx, col_name in enumerate(headers, 1):
            max_len = max(len(str(col_name)), combined[col_name].astype(str).str.len().max() if col_name in combined.columns else 0)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    ws2 = wb.create_sheet('Summary Report')
    ws2.column_dimensions['A'].width = 36
    ws2.column_dimensions['B'].width = 20

    ws2.append(['Pharmacy Data Migration Report'])
    ws2[ws2.max_row][0].font = Font(name='Arial', size=14, bold=True, color='1F4E79')
    ws2.append([f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'])
    ws2.append([f'Source file: {original_name}'])
    ws2.append([f'Supplier Name: {supplier_name}'])
    ws2.append([])

    total_in  = sum(s['input_rows'] for s in all_stats)
    total_out = sum(s['output_rows'] for s in all_stats)
    empty_rem = sum(s['empty_rows_removed'] for s in all_stats)

    def srow(label, value):
        ws2.append([label, value])
        for cell in ws2[ws2.max_row]:
            cell.font = Font(name='Arial', size=10)
            cell.border = THIN_BORDER

    srow('Total records received', total_in)
    srow('Empty rows removed', empty_rem)
    srow('Total medicines in output', total_out)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Routes ────────────────────────────────────────────────────────────────────

@app.errorhandler(Exception)
def handle_exception(e):
    import traceback
    return jsonify({'error': traceback.format_exc()}), 500


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/clean', methods=['POST'])
def clean():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    f = request.files['file']
    if not f.filename or not allowed_file(f.filename):
        return jsonify({'error': 'Unsupported file type. Upload .xlsx, .xls or .csv'}), 400

    patient_id_prefix = request.form.get('patient_id_prefix', '').strip().upper()
    if patient_id_prefix and not re.match(r'^[A-Z0-9]+$', patient_id_prefix):
        return jsonify({'error': 'Patient ID Prefix must be alphanumeric only.'}), 400
    if len(patient_id_prefix) > 20:
        return jsonify({'error': 'Patient ID Prefix must be 20 characters or fewer.'}), 400

    filename = secure_filename(f.filename)
    ext = filename.rsplit('.', 1)[1].lower()

    try:
        if ext == 'csv':
            raw = f.read()
            # Try common encodings used in Indian healthcare data
            for enc in ('utf-8', 'utf-8-sig', 'latin-1', 'cp1252', 'iso-8859-1'):
                try:
                    df = pd.read_csv(
                        io.BytesIO(raw), dtype=str,
                        keep_default_na=False, encoding=enc
                    )
                    break
                except (UnicodeDecodeError, Exception):
                    continue
            else:
                return jsonify({'error': 'Could not decode CSV file. Try saving it as UTF-8 from Excel.'}), 400
            df = df.replace('', pd.NA)
            sheets = {'Sheet1': df}
        else:
            xl = pd.ExcelFile(f)
            sheets = {}
            for sn in xl.sheet_names:
                sheets[sn] = xl.parse(sn, dtype=str).replace('', pd.NA)
    except Exception as e:
        return jsonify({'error': f'Could not read file: {str(e)}'}), 400

    all_dfs, all_stats = [], []
    try:
        for sn, df in sheets.items():
            if df.empty:
                continue
            cleaned_df, stats = process_dataframe(df, sheet_name=sn, patient_id_prefix=patient_id_prefix)
            if not cleaned_df.empty:
                all_dfs.append(cleaned_df)
            all_stats.append(stats)
    except Exception as e:
        import traceback
        return jsonify({'error': f'Processing error: {traceback.format_exc()}'}), 500

    if not all_stats:
        return jsonify({'error': 'No processable data found in the file'}), 400

    base_name = filename.rsplit('.', 1)[0]
    out_name = f'Cleaned_{base_name}.xlsx'
    buf = write_excel(all_dfs, all_stats, filename, patient_id_prefix=patient_id_prefix)

    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=out_name
    )


@app.route('/clean-pharmacy', methods=['POST'])
def clean_pharmacy():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    f = request.files['file']
    if not f.filename or not allowed_file(f.filename):
        return jsonify({'error': 'Unsupported file type. Upload .xlsx, .xls or .csv'}), 400

    supplier_name = request.form.get('supplier_name', '').strip()
    if not supplier_name:
        return jsonify({'error': 'Supplier name is required.'}), 400

    filename = secure_filename(f.filename)
    ext = filename.rsplit('.', 1)[1].lower()

    try:
        if ext == 'csv':
            raw = f.read()
            for enc in ('utf-8', 'utf-8-sig', 'latin-1', 'cp1252', 'iso-8859-1'):
                try:
                    df = pd.read_csv(io.BytesIO(raw), dtype=str, keep_default_na=False, encoding=enc)
                    break
                except Exception:
                    continue
            else:
                return jsonify({'error': 'Could not decode CSV. Try saving as UTF-8 from Excel.'}), 400
            df = df.replace('', pd.NA)
            sheets = {'Sheet1': df}
        else:
            xl = pd.ExcelFile(f)
            sheets = {sn: xl.parse(sn, dtype=str).replace('', pd.NA) for sn in xl.sheet_names}
    except Exception as e:
        return jsonify({'error': f'Could not read file: {str(e)}'}), 400

    all_dfs, all_stats = [], []
    try:
        for sn, df in sheets.items():
            if df.empty:
                continue
            cleaned_df, stats = process_pharmacy_dataframe(df, sheet_name=sn, supplier_name=supplier_name)
            if not cleaned_df.empty:
                all_dfs.append(cleaned_df)
            all_stats.append(stats)
    except Exception:
        import traceback
        return jsonify({'error': f'Processing error: {traceback.format_exc()}'}), 500

    if not all_dfs:
        return jsonify({'error': 'No medicine data found in the file.'}), 400

    base_name = filename.rsplit('.', 1)[0]
    out_name = f'Pharmacy_{base_name}.xlsx'
    buf = write_pharmacy_excel(all_dfs, all_stats, filename, supplier_name=supplier_name)

    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=out_name
    )


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=False)
